from odoo import models, fields, _
from odoo.exceptions import UserError
import base64
from io import BytesIO
import openpyxl
import msoffcrypto

class ExcelImportWizard(models.TransientModel):
    _name = 'excel.import.wizard'

    excel_file = fields.Binary(string='Excel File', required=True)
    password = fields.Char(string='Password', default='IPACRCI')


    def import_excel_data(self):
        if not self.excel_file:
            raise UserError(_("Please upload an Excel file."))

        try:
            file_content = base64.b64decode(self.excel_file)
            decrypted_file = BytesIO()
            office_file = msoffcrypto.OfficeFile(BytesIO(file_content))
            office_file.load_key(password=self.password)
            office_file.decrypt(decrypted_file)
            decrypted_file.seek(0)

            # Debugging: print a message before reading the file
            print("Déchiffrement réussi, tentative de lecture du fichier Excel...")

            # Try loading the workbook with different parameters
            try:
                excel_workbook = openpyxl.load_workbook(filename=decrypted_file, keep_links=False)
            except TypeError as e:
                print(f"Erreur de type lors de la lecture du fichier Excel : {e}")
                excel_workbook = openpyxl.load_workbook(filename=decrypted_file)

            excel_sheet = excel_workbook.active
            excel_data = excel_sheet.iter_rows(values_only=True)
        except Exception as e:
            raise UserError(_("Failed to open Excel file: %s" % e))

        for row in excel_data:
            self.env['excel.imported.data'].create({
                'field1': row[0],
                'field2': row[1],
            })

        raise UserError(_("Excel data imported successfully."))

class ExcelImportedData(models.Model):
    _name = 'excel.imported.data'
    _description = 'Excel Imported Data'

    field1 = fields.Char(string='Field 1')
    field2 = fields.Char(string='Field 2')
