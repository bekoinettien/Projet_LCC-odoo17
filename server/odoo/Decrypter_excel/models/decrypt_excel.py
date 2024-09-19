# models/decrypt_excel.py
from odoo import models, fields, api
from openpyxl import load_workbook
from Cryptodome.Cipher import AES
import base64
import tempfile
import binascii

class ExcelDecryptor(models.Model):
    _name = 'excel.decryptor'
    _description = 'Decrypt Excel Files'

    name = fields.Char(string='Name')
    file = fields.Binary(string='File')
    file_name = fields.Char(string='File Name')
    password = fields.Char(string='Password', invisible=True)
    decrypted_content = fields.Text(string='Decrypted Content', readonly=True)

    def action_decrypt_and_view(self):
        for record in self:
            if record.file and record.password:
                file_path = self._save_file(record.file, record.file_name)
                workbook = self.decrypt_file(file_path, record.password)
                content = self._get_workbook_content(workbook)
                record.decrypted_content = content

    def _save_file(self, file_content, file_name):
        with tempfile.NamedTemporaryFile(delete=False, suffix=file_name) as tmp_file:
            tmp_file.write(binascii.a2b_base64(file_content))
            return tmp_file.name

    def decrypt_file(self, file_path, password):
        workbook = load_workbook(filename=file_path, read_only=False)

        def decrypt(content, password):
            key = password.ljust(32)[:32].encode('utf-8')
            cipher = AES.new(key, AES.MODE_ECB)
            decrypted = cipher.decrypt(base64.b64decode(content))
            return decrypted.strip()

        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, str):
                        try:
                            decrypted_value = decrypt(cell.value, password)
                            cell.value = decrypted_value
                        except Exception as e:
                            pass

        return workbook

    def _get_workbook_content(self, workbook):
        content = ''
        for sheet in workbook.worksheets:
            content += f'Sheet: {sheet.title}\n'
            for row in sheet.iter_rows(values_only=True):
                content += '\t'.join([str(cell) if cell is not None else '' for cell in row]) + '\n'
        return content
